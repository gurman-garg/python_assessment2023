#Version 3: GUI Version of Cafe Click & Collect

#importing tkinter, messagebox
#importing openpyxl so excel file can be linked to python file
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook

#defining path to excel file
excel_file = "Excel/python_cafe_accounts.xlsx"

def register_account(add_username, add_password, excel_file, text_widget):
    if int(register_age.get()) in range(13, 19):
        #loading the workbook
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        #checking if the username and password combination exists
        existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]

        if (add_username, add_password) not in existing_data:
            #add the username and password to the excel file
            sheet.append([add_username, add_password])

            #saving the changes
            workbook.save(excel_file)
            
            #telling user that the account has been registered and opening the ordering window
            text_widget.delete("1.0", END)
            text_widget.insert(END, "Account Registered!")
            open_ordering_window()
        else:
            #if the username and password combination is in the existing data in the excel sheet, ask user to sign in instead
            text_widget.delete("1.0", END)
            text_widget.insert(END, "Username and Password already exist! Please Sign In instead.")
    else:
        #if the user is not in the age range, tell them they cannot register for the click & collect
        text_widget.delete("1.0", END)
        text_widget.insert(END, "You are not the right age to use this program.")        

#the function that prints the order in a messagebox after user has confirmed the order
def print_order():
    #defining these variables as global so they can be used in this function
    global item1_selected, item2_selected, item3_selected, item4_selected, item5_selected

    order_list = []
    total_cost = 0

    #retrieving the updated values of the IntVar variables
    item1_value = item1_selected.get()
    item2_value = item2_selected.get()
    item3_value = item3_selected.get()
    item4_value = item4_selected.get()
    item5_value = item5_selected.get()

    if item1_value == 1:  #check if the checkbox is selected
        order_list.append("Hash Brown - $1.50")
        total_cost += 1.50

    if item2_value == 1:
        order_list.append("Cookie - $3.50")
        total_cost += 3.50

    if item3_value == 1:
        order_list.append("Nachos - $4.50")
        total_cost += 4.50

    if item4_value == 1:
        order_list.append("Garlic Bread - $2.50")
        total_cost += 2.50

    if item5_value == 1:
        order_list.append("Burger - $5.00")
        total_cost += 5.00

    order_text = "\n".join(order_list)
    total_text = f"Total Cost: ${total_cost:.2f}"
    messagebox.showinfo("Order", f"Order:\n{order_text}\n{total_text}\n\nPlease pick up your order from the cafe.")

#sign in function that checks if the username and password exist and are correct
def sign_in():
    #load the workbook
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    #checking if username and password combination exists
    existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]

    username = sign_username.get()
    password = sign_password.get()
    
    #if username and password are correct and in existing data, open ordering window
    if (str(username), str(password)) in existing_data:
        text1.delete("1.0", END)
        text1.insert(END, "Login Successful!")
        open_ordering_window()
    else:
        text1.delete("1.0", END)
        text1.insert(END, "Username or Password incorrect. Try Again.")

def open_ordering_window():
    #declaring variables as global so they can be accessed in this function
    global item1_selected, item2_selected, item3_selected, item4_selected, item5_selected

    #making a new window called "ordering_window" as a toplevel window
    ## A toplevel window is a separate window that can be created on top of the main window to display additional content without affecting the main window
    ordering_window = Toplevel(root)
    ordering_window.title("BDSC Cafe Click & Collect")
    ordering_window.geometry("250x250")

    info_label = Label(ordering_window, text="BDSC Cafe Click & Collect")
    info_label.grid(column=0, row=0)
    
    #making checkbutton variables IntVar variables
    item1_selected = IntVar()
    item2_selected = IntVar()
    item3_selected = IntVar()
    item4_selected = IntVar()
    item5_selected = IntVar()

    #checkbuttons (sticky used so they are aligned)
    item1_checkbox = Checkbutton(ordering_window, text="Hash Brown      $1.50", variable=item1_selected)
    item1_checkbox.grid(column=0, row=1, sticky = W)

    item2_checkbox = Checkbutton(ordering_window, text="Cookie          $3.50", variable=item2_selected)
    item2_checkbox.grid(column=0, row=2, sticky = W)

    item3_checkbox = Checkbutton(ordering_window, text="Nachos          $4.50", variable=item3_selected)
    item3_checkbox.grid(column=0, row=3, sticky = W)

    item4_checkbox = Checkbutton(ordering_window, text="Garlic Bread    $2.50", variable=item4_selected)
    item4_checkbox.grid(column=0, row=4, sticky = W)

    item5_checkbox = Checkbutton(ordering_window, text="Burger          $5.00", variable=item5_selected)
    item5_checkbox.grid(column=0, row=5, sticky = W)

    order_button = Button(ordering_window, text="Confirm Order", command=print_order)
    order_button.grid(column=0, row=6)


#main code that starts the program, it's also the register/sign in window (as required by the brief)
root = Tk()
root.title("BDSC Cafe Sign In / Register")
root.geometry("400x350")
        
intro_label = Label(root, text="BDSC Cafe Click & Collect Sign In / Register")
intro_label.grid(column=1, row=0)
        
sign_label = Label(root, text="Sign In")
sign_label.grid(column=1, row=1)
        
sign_user_label = Label(root, text="Username: ")
sign_user_label.grid(column=0, row=2)
        
sign_pw_label = Label(root, text="Password: ")
sign_pw_label.grid(column=0, row=3)
        
sign_username = Entry(root)
sign_username.grid(column=1, row=2)
        
sign_password = Entry(root)
sign_password.grid(column=1, row=3)
        
signin_button = Button(root, text="Sign In", command=sign_in)
signin_button.grid(column=1, row=4)
        
register_label = Label(root, text="Register")
register_label.grid(column=1, row=5)
        
register_user_label = Label(root, text="Username:")
register_user_label.grid(column=0, row=6)
        
register_pw_label = Label(root, text="Password:")
register_pw_label.grid(column=0, row=7)
        
register_age_label = Label(root, text="Age:")
register_age_label.grid(column=0, row=8)
        
register_username = Entry(root)
register_username.grid(column=1, row=6)
        
register_password = Entry(root)
register_password.grid(column=1, row=7)
        
register_age = Entry(root)
register_age.grid(column=1, row=8)
        
register_button = Button(root, text="Register", command=lambda: register_account(register_username.get(), register_password.get(), excel_file, text1))
register_button.grid(column=1, row=9)
        
text1 = Text(root, width=28, height=2)
text1.grid(column=1, row=10)
        
root.mainloop()