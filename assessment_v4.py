#Version 4: Improved GUI Version of Cafe Click & Collect

#importing tkinter, messagebox
#importing openpyxl so excel file can be linked to python file
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook

#defining path to excel file
excel_file = "Excel/python_cafe_accounts.xlsx"

#dictionary to store name of menu items and their prices
item_prices = {
    "Hash Brown": 1.50,
    "Cookie": 3.50,
    "Nachos": 4.50,
    "Garlic Bread": 2.50,
    "Burger": 5.00
}

item_quantities = {}  #defining item_quantities as a global variable

def register_account(add_username, add_password, excel_file, text_widget):
    if int(register_age.get()) in range(13, 19):
        workbook = load_workbook(excel_file)
        sheet = workbook["Sheet1"]
        
        #checking if the username and password combination exists
        existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]
        
        if (add_username, add_password) not in existing_data:
            #add the username and password to the excel file
            sheet.append([add_username, add_password])
            
            #saving the changes
            workbook.save(excel_file)
            
            #telling user that the account has been registered and opening the ordering window
            text_widget.delete("1.0", END)
            text_widget.insert(END, "Account Registered!")
            open_ordering_window(add_username)
        else:
            #if the username and password combination is in the existing data in the excel sheet, ask user to sign in instead
            text_widget.delete("1.0", END)
            text_widget.insert(END, "Username and Password already exist! Please Sign In instead.")
    else:
        #if the user is not in the age range, tell them they cannot register for the click & collect
        text_widget.delete("1.0", END)
        text_widget.insert(END, "You are not the right age to use this program.")

#sign in function that checks if the username and password exist and are correct
def sign_in():
    username = sign_username.get()
    password = sign_password.get()

    workbook = load_workbook(excel_file)
    sheet = workbook["Sheet1"]
    
    #checking if username and password combination exists
    existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]

    #if username and password are correct and in existing data, open ordering window
    if (username, password) in existing_data:
        text1.delete("1.0", END)
        text1.insert(END, "Login Successful!")
        open_ordering_window(username)
    else:
        text1.delete("1.0", END)
        text1.insert(END, "Username or Password incorrect. Try Again.")

#the function that prints the order in a messagebox after user has confirmed the order
def print_order(username):
    #defining variable as global so it can be used in this function
    global item_quantities
    order_list = []
    total_cost = 0
    
    #item represents the item name and quantity_widget represents the corresponding Spinbox widget    
    for item, quantity_widget in item_quantities.items():
        
        # Get the quantity selected for the item from the Spinbox widget and adding ordered items with quantity to order list and getting total cost
        quantity = int(quantity_widget.get())
        if quantity > 0:
            price_per_item = item_prices[item]
            total_price = quantity * price_per_item
            order_list.append(f"{item} - ${price_per_item:.2f} (x{quantity})")
            total_cost += total_price
    
    #printing to messagebox
    order_text = "\n".join(order_list)
    total_text = f"Total Cost: ${total_cost:.2f}"
    messagebox.showinfo("Order", f"Order:\n{order_text}\n\n{total_text}\n\nPlease pick up your order from the cafe.")

    #adding order for the account to sheet 2 of excel sheet
    workbook = load_workbook(excel_file)
    sheet2 = workbook["Sheet2"]
    sheet2.append([username, order_text, total_cost])
    workbook.save(excel_file)

def open_ordering_window(username):
    #defining global variables as global and create new window for ordering
    global item_quantities
    ordering_window = Toplevel(root)
    ordering_window.title("BDSC Cafe Click & Collect")
    ordering_window.geometry("350x250")
    info_label = Label(ordering_window, text="BDSC Cafe Click & Collect")
    info_label.grid(column=0, row=0)

    item_quantities = {}

    # Iterate over the item_prices dictionary using enumerate to get the index and item/price pair (enumerate lets you go through each item in a list and know its position in the list)
    for i, (item, price) in enumerate(item_prices.items()):
        item_label = Label(ordering_window, text=f"{item} - ${price:.2f}")
        item_label.grid(column=0, row=i+1)

        item_quantity = Spinbox(ordering_window, from_=0, to=10, width=10)
        item_quantity.grid(column=1, row=i+1)

        item_quantities[item] = item_quantity

    order_button = Button(ordering_window, text="Confirm Order", command=lambda: print_order(username))
    #len counts the number of items in a list
    order_button.grid(column=0, row=len(item_prices) + 1)


#main code that starts the program, it's also the register/sign in window (as required by the brief)
root = Tk()
root.title("BDSC Cafe Sign In / Register")
root.geometry("400x350")

intro_label = Label(root, text="BDSC Cafe Click & Collect Sign In / Register")
intro_label.grid(column=1, row=0)

sign_label = Label(root, text="Sign In")
sign_label.grid(column=1, row=1)

sign_user_label = Label(root, text="Username: ")
sign_user_label.grid(column=0, row=2)

sign_pw_label = Label(root, text="Password: ")
sign_pw_label.grid(column=0, row=3)

sign_username = Entry(root)
sign_username.grid(column=1, row=2)

sign_password = Entry(root, show="*")
sign_password.grid(column=1, row=3)

signin_button = Button(root, text="Sign In", command=sign_in)
signin_button.grid(column=1, row=4)

register_label = Label(root, text="Register")
register_label.grid(column=1, row=5)

register_user_label = Label(root, text="Username:")
register_user_label.grid(column=0, row=6)

register_pw_label = Label(root, text="Password:")
register_pw_label.grid(column=0, row=7)

register_age_label = Label(root, text="Age:")
register_age_label.grid(column=0, row=8)

register_username = Entry(root)
register_username.grid(column=1, row=6)

register_password = Entry(root)
register_password.grid(column=1, row=7)

register_age = Entry(root)
register_age.grid(column=1, row=8)

register_button = Button(root, text="Register", command=lambda: register_account(register_username.get(), register_password.get(), excel_file, text1))
register_button.grid(column=1, row=9)

text1 = Text(root, width=28, height=2)
text1.grid(column=1, row=10)

root.mainloop()