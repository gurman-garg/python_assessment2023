#Version 5: Improved GUI Version of Cafe Click & Collect with Images

#importing tkinter, messagebox, openpyxl (for excel file), and PIL (for images)
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
from PIL import ImageTk, Image 

#defining the path to the excel file
excel_file = "Excel/python_cafe_accounts.xlsx"

#dictionary to store name of menu items and their prices
item_prices = {
    "Hash Brown": 1.50,
    "Cookie": 3.50,
    "Nachos": 4.50,
    "Garlic Bread": 2.50,
    "Burger": 5.00
}

item_quantities = {} #defining item_quantities as a global variable

#function to toggle password visibility in entry widgets
def toggle_password_visibility(entry_widget, toggle_button):
    if entry_widget['show'] == "*":
        entry_widget['show'] = ""
        toggle_button.config(text="Hide")
    else:
        entry_widget['show'] = "*"
        toggle_button.config(text="Show")

#function to register a new account
def register_account(add_username, add_password, confirm_password, excel_file, text_widget):
    
    #check if age entered is in range
    register_age_value = register_age.get()
    if register_age_value.isdigit() and int(register_age_value) in range(13, 19):
        
        #loading excel workbook/sheet
        workbook = load_workbook(excel_file)
        sheet = workbook["Sheet1"]
        
        #checking if the username and password combination exists
        existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]

        #adding account to excel sheet if input in password and confirm password fields match
        if (add_username, add_password) not in existing_data:
            if add_password == confirm_password:
                sheet.append([add_username, add_password])
                
                #saving the changes
                workbook.save(excel_file)
                
                #telling user that the account has been registered and opening the ordering window
                text_widget.delete("1.0", END)
                text_widget.insert(END, "Account Registered!")
                open_ordering_window(add_username)
            else:
                #when the passwords in the password and confirm password fields do not match
                text_widget.delete("1.0", END)
                text_widget.insert(END, "Passwords do not match. Please try again.")
        else:
            #if the username and password combination is in the existing data in the excel sheet, ask user to sign in instead
            text_widget.delete("1.0", END)
            text_widget.insert(END, "Username and Password already exist! Please Sign In instead.")
    else:
        #if the user is not in the age range, tell them they cannot register for the click & collect
        text_widget.delete("1.0", END)
        text_widget.insert(END, "Invalid age. This program is only for students.")

#function for user sign-in
def sign_in():
    username = sign_username.get()
    password = sign_password.get()
    
    #loading the excel sheet/workbook
    workbook = load_workbook(excel_file)
    sheet = workbook["Sheet1"]
    
    #checking if username and password combination exists
    existing_data = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True)]

    #if username and password are correct and in existing data, open ordering window
    if (username, password) in existing_data:
        text1.delete("1.0", END)
        text1.insert(END, "Login Successful!")
        open_ordering_window(username)
    else:
        text1.delete("1.0", END)
        text1.insert(END, "Username or Password incorrect. Try Again.")

#function to print the order details in messagebox
def print_order(username):
    
    #defining variable as global so it can be used in this function
    global item_quantities
    order_list = []
    total_cost = 0
    
    #item represents the item name and quantity_widget represents the corresponding Spinbox widget 
    for item, quantity_widget in item_quantities.items():
        
        #get the quantity selected for the item from the Spinbox widget and adding ordered items with quantity to order list and getting total cost        
        quantity = int(quantity_widget.get())
        if quantity > 0:
            price_per_item = item_prices[item]
            total_price = quantity * price_per_item
            order_list.append(f"{item} - ${price_per_item:.2f} (x{quantity})")
            total_cost += total_price

    order_text = "\n".join(order_list)
    total_text = f"Total Cost: ${total_cost:.2f}"

    #adding order for the account to sheet 2 of excel sheet
    workbook = load_workbook(excel_file)
    sheet = workbook["Sheet2"]

    last_row = sheet.max_row + 1
    sheet[f"A{last_row}"] = username
    sheet[f"B{last_row}"] = order_text
    sheet[f"C{last_row}"] = total_cost
    
    #saving changes in excel file
    workbook.save(excel_file)

    #printing order and total cost in messagebox
    messagebox.showinfo("Order", f"Order:\n{order_text}\n\n{total_text}\n\nPlease pick up your order from the cafe.")

#function to open the ordering window
def open_ordering_window(username):
    
    #defining global variables as global and create new window for ordering
    global item_quantities
    ordering_window = Toplevel(root)
    ordering_window.title("BDSC Cafe Click & Collect")
    ordering_window.geometry("350x250")
    info_label = Label(ordering_window, text="BDSC Cafe Click & Collect")
    info_label.grid(column=0, row=0)

    item_quantities = {}
   
    #enumerate lets you go through each item in a list and know its position in the list
    for i, (item, price) in enumerate(item_prices.items()):
        item_label = Label(ordering_window, text=f"{item} - ${price:.2f}")
        item_label.grid(column=0, row=i+1)

        item_quantity = Spinbox(ordering_window, from_=0, to=10, width=10)
        item_quantity.grid(column=1, row=i+1)

        item_quantities[item] = item_quantity

    order_button = Button(ordering_window, text="Confirm Order", command=lambda: [print_order(username)])
    #len counts the number of items in a list
    order_button.grid(column=0, row=len(item_prices) + 1)

#creating main application window, it's also the register/sign in window (as required by the brief)
root = Tk()
root.title("BDSC Cafe Sign In / Register")
root.geometry("500x400")

#creating and placing various GUI elements/widgets in the main window
intro_label = Label(root, text="BDSC Cafe Click & Collect Sign In / Register")
intro_label.grid(column=1, row=0)

##adding image to window
profile_img = ImageTk.PhotoImage(Image.open("Images/profile4.png"))
img_label = Label(root, image=profile_img)
img_label.grid(column = 2, row = 2)

sign_label = Label(root, text="Sign In")
sign_label.grid(column=1, row=1)

sign_user_label = Label(root, text="Username: ")
sign_user_label.grid(column=0, row=2)

sign_pw_label = Label(root, text="Password: ")
sign_pw_label.grid(column=0, row=3)

sign_username = Entry(root)
sign_username.grid(column=1, row=2)

sign_password = Entry(root, show="*")
sign_password.grid(column=1, row=3)

sign_toggle_button = Button(root, text="Show", command=lambda: toggle_password_visibility(sign_password, sign_toggle_button))
sign_toggle_button.grid(column=2, row=3)

signin_button = Button(root, text="Sign In", command=sign_in)
signin_button.grid(column=1, row=4)

register_label = Label(root, text="Register")
register_label.grid(column=1, row=5)

register_user_label = Label(root, text="Username:")
register_user_label.grid(column=0, row=6)

register_pw_label = Label(root, text="Password:")
register_pw_label.grid(column=0, row=7)

confirm_pw_label = Label(root, text="Confirm Password:")
confirm_pw_label.grid(column=0, row=8)

register_username = Entry(root)
register_username.grid(column=1, row=6)
    
register_password = Entry(root, show="*")
register_password.grid(column=1, row=7)
    
register_confirm_password = Entry(root, show="*")
register_confirm_password.grid(column=1, row=8)
    
register_toggle_button = Button(root, text="Show", command=lambda: toggle_password_visibility(register_password, register_toggle_button))
register_toggle_button.grid(column=2, row=7)
    
confirm_toggle_button = Button(root, text="Show", command=lambda: toggle_password_visibility(register_confirm_password, confirm_toggle_button))
confirm_toggle_button.grid(column=2, row=8)
    
register_age_label = Label(root, text="Age:")
register_age_label.grid(column=0, row=9)
    
register_age = Entry(root)
register_age.grid(column=1, row=9)
    
register_button = Button(root, text="Register", command=lambda: register_account(register_username.get(), register_password.get(), register_confirm_password.get(), excel_file, text1))
register_button.grid(column=1, row=10)
    
text1 = Text(root, width=28, height=2)
text1.grid(column=1, row=11)
    
root.mainloop()